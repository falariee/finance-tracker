"""
Expense Service - Business logic for expense management
"""

import json
import csv
from models.expense import Expense


class ExpenseService:
    def __init__(self):
        self.trip = None
        self.expenses = []
    
    def set_trip(self, trip):
        """Set the current trip"""
        self.trip = trip
    
    def add_expense(self, expense):
        """Add an expense to the trip"""
        self.expenses.append(expense)
    
    def get_all_expenses(self):
        """Get all expenses"""
        return self.expenses
    
    def get_expenses_by_category(self, category):
        """Get expenses filtered by category"""
        return [exp for exp in self.expenses if exp.category == category]
    
    def get_expenses_by_person(self, person_name):
        """Get expenses paid by a specific person"""
        return [exp for exp in self.expenses if exp.paid_by == person_name]
    
    def get_total_expenses(self):
        """Calculate total expenses"""
        return sum(exp.amount for exp in self.expenses)
    
    def get_category_totals(self):
        """Get total expenses by category"""
        categories = {}
        for exp in self.expenses:
            if exp.category not in categories:
                categories[exp.category] = 0
            categories[exp.category] += exp.amount
        return categories
    
    def get_person_totals(self):
        """Get total expenses by person"""
        people = {}
        for exp in self.expenses:
            if exp.paid_by not in people:
                people[exp.paid_by] = 0
            people[exp.paid_by] += exp.amount
        return people
    
    def export_to_json(self, filename):
        """Export expenses to JSON"""
        data = {
            "trip": self.trip.to_dict() if self.trip else None,
            "expenses": [exp.to_dict() for exp in self.expenses]
        }
        with open(filename, 'w') as f:
            json.dump(data, f, indent=2)
    
    def export_to_csv(self, filename):
        """Export expenses to CSV"""
        if not self.expenses:
            return
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['date', 'description', 'amount', 'currency', 
                         'converted_amount', 'trip_currency', 'category', 'paid_by']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            
            writer.writeheader()
            for exp in self.expenses:
                row = {
                    'date': exp.date,
                    'description': exp.description,
                    'amount': exp.amount,
                    'currency': exp.currency,
                    'converted_amount': exp.amount if self.trip and exp.currency == self.trip.currency else '',
                    'trip_currency': self.trip.currency if self.trip else '',
                    'category': exp.category,
                    'paid_by': exp.paid_by
                }
                writer.writerow(row)
    
    def export_to_excel(self, filename):
        """Export expenses to Excel with summary"""
        if not self.expenses:
            return
        
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Expenses"
        
        # Header style
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add headers
        headers = ['Date', 'Description', 'Amount', 'Currency', 'Category', 'Paid By']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add expenses
        for row_idx, exp in enumerate(self.expenses, 2):
            ws.cell(row=row_idx, column=1, value=exp.date).border = border
            ws.cell(row=row_idx, column=2, value=exp.description).border = border
            ws.cell(row=row_idx, column=3, value=exp.amount).border = border
            ws.cell(row=row_idx, column=4, value=exp.currency).border = border
            ws.cell(row=row_idx, column=5, value=exp.category).border = border
            ws.cell(row=row_idx, column=6, value=exp.paid_by).border = border
        
        # Add summary section
        summary_row = len(self.expenses) + 3
        
        # Summary header
        ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        summary_row += 2
        
        # Calculate totals by currency
        currency_totals = {}
        for exp in self.expenses:
            if exp.currency not in currency_totals:
                currency_totals[exp.currency] = 0
            currency_totals[exp.currency] += exp.amount
        
        # Add currency breakdown
        ws.cell(row=summary_row, column=1, value="Total by Currency:").font = Font(bold=True)
        summary_row += 1
        
        for currency, total in sorted(currency_totals.items()):
            ws.cell(row=summary_row, column=1, value=currency)
            ws.cell(row=summary_row, column=2, value=total).number_format = '#,##0.00'
            summary_row += 1
        
        summary_row += 1
        
        # Category breakdown
        ws.cell(row=summary_row, column=1, value="Total by Category:").font = Font(bold=True)
        summary_row += 1
        
        category_totals = self.get_category_totals()
        for category, total in sorted(category_totals.items()):
            ws.cell(row=summary_row, column=1, value=category.capitalize())
            ws.cell(row=summary_row, column=2, value=total).number_format = '#,##0.00'
            if self.trip:
                ws.cell(row=summary_row, column=3, value=self.trip.currency)
            summary_row += 1
        
        summary_row += 1
        
        # Person totals
        ws.cell(row=summary_row, column=1, value="Total by Person:").font = Font(bold=True)
        summary_row += 1
        
        person_totals = self.get_person_totals()
        for person, total in sorted(person_totals.items()):
            ws.cell(row=summary_row, column=1, value=person)
            ws.cell(row=summary_row, column=2, value=total).number_format = '#,##0.00'
            if self.trip:
                ws.cell(row=summary_row, column=3, value=self.trip.currency)
            summary_row += 1
        
        # Auto-adjust column widths
        for col in range(1, 7):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(filename)
    
    def import_from_json(self, filename):
        """Import expenses from JSON"""
        with open(filename, 'r') as f:
            data = json.load(f)
        
        if data.get("trip"):
            from models.trip import Trip
            self.trip = Trip.from_dict(data["trip"])
        
        self.expenses = [Expense.from_dict(exp) for exp in data.get("expenses", [])]
